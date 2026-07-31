"""
UCREW -- Commercial & Industrial Energy Efficiency Incentives Database
Scope: set by ENABLED_STATES below. Currently Utah (UT) only; add codes to grow.
Supported states: Utah (UT), Montana (MT), Idaho (ID), Nevada (NV)

Scope: commercial & industrial (C&I) facility incentives only -- residential,
multifamily, and new-home/builder programs are intentionally excluded.

Run manually:  python fetch_incentives.py
Scheduled:     GitHub Actions (.github/workflows/update-incentives.yml) runs daily
               and publishes the site/ folder to GitHub Pages.

Outputs:
  incentives.db    -- SQLite database (all fields including detail content)
  incentives.xlsx  -- Excel workbook (summary columns + Details sheet)
  incentives.html  -- Browser-viewable sortable table with clickable detail modals
  site/            -- Static site published to GitHub Pages (index.html + downloads)
  run_log.txt      -- Timestamped run history
"""

import sqlite3
import importlib
import json
import shutil
import base64
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from collections import Counter

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "incentives.db"
XLSX_PATH = BASE_DIR / "incentives.xlsx"
HTML_PATH = BASE_DIR / "incentives.html"
LOG_PATH = BASE_DIR / "run_log.txt"
# GitHub Pages publishes this folder (Settings -> Pages -> Source: GitHub Actions).
SITE_DIR = BASE_DIR / "site"
# UCREW brand logo (vendored in-repo so CI builds have no external dependency).
LOGO_PATH = BASE_DIR / "assets" / "ucrew-logo.svg"

# UCREW brand palette (derived from the logo red #be0000). Used to theme the site.
BRAND = "#be0000"        # UCREW red -- primary
BRAND_DARK = "#8a0000"   # hover / gradient bottom
BRAND_DARKER = "#6b0000" # deepest shade


def _logo_data_uri():
    """Return the UCREW logo as a base64 data URI so the HTML stays self-contained
    (works when double-clicked and when served from GitHub Pages, no asset copy)."""
    try:
        raw = LOGO_PATH.read_bytes()
    except OSError:
        return ""
    b64 = base64.b64encode(raw).decode("ascii")
    return "data:image/svg+xml;base64," + b64


STATE_COLORS = {
    "UT": "FF4E79A7",
    "MT": "FF59A14F",
    "ID": "FFF28E2B",
    "NV": "FFE15759",
}

# Canonical ordering + display names for every state the scanner *can* cover.
ALL_STATES = ["UT", "MT", "ID", "NV"]
STATE_NAMES = {"UT": "Utah", "MT": "Montana", "ID": "Idaho", "NV": "Nevada"}

# States the scanner is *currently* scoped to. Start narrow (Utah only) and grow
# by adding codes here as each state is perfected -- e.g. ["UT", "ID"]. This one
# constant drives the scrapers that run, the rows kept, and all site/Excel labels.
ENABLED_STATES = ["UT"]

# Summary columns shown in Excel main sheets and HTML table
COLUMNS = [
    "State", "Program Name", "Administrator", "Sector", "Incentive Type",
    "Technology", "Incentive Value", "Max Benefit", "Eligible Recipients",
    "Expiration Date", "Application URL", "Last Scraped", "Status", "Notes",
]

# Detail columns stored in SQLite and shown in modal / Excel Details sheet.
# The _ -prefixed structured fields hold the actual numbers used in calculations.
DETAIL_COLUMNS = [
    "_incentive_rate", "_rebate_tiers", "_unit_cap", "_baseline", "_min_project",
    "_implementation", "_methodology", "_example",
]

ALL_COLUMNS = COLUMNS + DETAIL_COLUMNS


# ---------------------------------------------------------------------------
# Equipment classification (UCREW audit categories).
#
# ONE keyword vocabulary drives two things:
#   1. build-time: every program is auto-tagged with the equipment categories
#      its text mentions (name + technology + methodology + example + notes).
#   2. run-time: the "AR Finder" box in the site maps a typed assessment
#      recommendation (e.g. "Install VFD on compressors") to the same
#      categories, then surfaces the programs tagged with them.
#
# Matching is case-insensitive and word-boundary aware: a keyword matches only
# at the start of a word, but trailing letters are allowed so the lowercase root
# "compressor" also matches "compressors". This prevents false hits like "led"
# inside "controlled". Recall is favored over precision -- a discovery aid, not a
# billing engine.
# ---------------------------------------------------------------------------
EQUIPMENT_KEYWORDS = {
    "Lighting": [
        "lighting", "light fixture", "led", "lamp", "luminaire", "high bay",
        "high-bay", "troffer", "t8", "t5", "fluorescent", "daylighting",
        "occupancy sensor", "exit sign", "delamping",
    ],
    "Compressed Air": [
        "compressed air", "compressor", "air compressor", "vsd compressor",
        "air dryer", "air leak", "compressed-air",
    ],
    "HVAC": [
        "hvac", "air conditioning", "air conditioner", "rooftop unit", "rtu",
        "chiller", "cooling", "furnace", "heat pump", "thermostat", "economizer",
        "ventilation", "make-up air", "makeup air", "space heating", "packaged unit",
        "mini split", "mini-split", "vrf",
    ],
    "Boilers & Steam": [
        "boiler", "steam", "condensate", "steam trap", "burner", "hot water heating",
        "linkageless", "combustion",
    ],
    "Motors & Drives": [
        "vfd", "variable frequency drive", "variable-frequency", "vsd",
        "variable speed", "adjustable speed", "premium efficiency motor",
        "efficient motor", "motor", "drive", "ecm",
    ],
    "Pumps": [
        "pump", "pumping",
    ],
    "Fans": [
        "fan", "exhaust fan", "ec motor", "ceiling fan",
    ],
    "Refrigeration": [
        "refrigeration", "refrigerated", "walk-in cooler", "walk-in freezer",
        "walk in cooler", "cooler", "freezer", "evaporator", "night cover",
        "anti-sweat", "strip curtain", "case lighting",
    ],
    "Process Heating": [
        "process heat", "oven", "kiln", "industrial dryer", "heat recovery",
        "waste heat", "process load",
    ],
    "Building Envelope": [
        "insulation", "weatherization", "envelope", "window", "roof", "air sealing",
        "pipe insulation", "cool roof", "door seal",
    ],
    "Water Heating": [
        "water heater", "water heating", "domestic hot water", "dhw",
        "heat pump water heater", "hpwh", "tankless",
    ],
    "Controls / EMS": [
        "controls", "control system", "ems", "bms", "building automation",
        "energy management system", "setback", "network lighting control",
        "sensor", "retrocommissioning", "retro-commissioning", "commissioning",
    ],
    "Irrigation": [
        "irrigation", "sprinkler", "center pivot", "agricultural pump",
        "wire-to-water", "scientific irrigation scheduling",
    ],
    "Custom / Whole Facility": [
        "custom", "multiple technolog", "whole building", "whole-building",
        "whole facility", "comprehensive", "strategic energy management",
        "energy audit", "new construction", "any measure", "all measures",
    ],
}

# Programs tagged with any of these categories are treated as broad measures
# that plausibly cover almost any recommendation, so the AR Finder always
# surfaces them (flagged separately as "custom / whole-facility").
CUSTOM_CATEGORIES = {"Custom / Whole Facility"}


def _kw_hit(keyword, text):
    """True if ``keyword`` appears at a word boundary in ``text`` (trailing
    letters allowed, so "compressor" also matches "compressors")."""
    return re.search(r"\b" + re.escape(keyword), text) is not None


def classify_equipment(text):
    """Return the sorted list of equipment categories whose keywords appear in
    ``text`` (case-insensitive, word-boundary aware)."""
    t = (text or "").lower()
    hits = [cat for cat, kws in EQUIPMENT_KEYWORDS.items()
            if any(_kw_hit(kw, t) for kw in kws)]
    return sorted(hits)



def main():
    print("\n" + "=" * 60)
    print("  Energy Efficiency Incentives DB -- " + str(date.today()))
    print("=" * 60 + "\n")

    all_rows = _fetch_all_sources()
    all_rows = _deduplicate(all_rows)
    all_rows = _auto_expire(all_rows)

    print("\nTotal programs collected: " + str(len(all_rows)))
    _print_summary(all_rows)

    _write_sqlite(all_rows)
    _write_excel(all_rows)
    _write_html(all_rows)
    _stage_site()
    _write_log(len(all_rows))

    print("\nDone. Files written:")
    print("  " + str(XLSX_PATH))
    print("  " + str(HTML_PATH))
    print("  " + str(DB_PATH))
    print("  " + str(SITE_DIR) + " (published to GitHub Pages)")


def _stage_site():
    """Assemble the folder GitHub Pages serves: index.html + downloadable data files."""
    SITE_DIR.mkdir(exist_ok=True)
    shutil.copy2(HTML_PATH, SITE_DIR / "index.html")
    if XLSX_PATH.exists():
        shutil.copy2(XLSX_PATH, SITE_DIR / "incentives.xlsx")
    if DB_PATH.exists():
        shutil.copy2(DB_PATH, SITE_DIR / "incentives.db")
    print("Site staged -> " + str(SITE_DIR / "index.html"))


def _fetch_all_sources():
    rows = []
    print("Fetching data sources (states: " + ", ".join(ENABLED_STATES) + ")...\n")

    # (label, module, func, state). state=None means the source is multi-state
    # and is scoped by passing ENABLED_STATES; single-state sources are skipped
    # entirely when their state is not enabled.
    sources = [
        ("Rocky Mountain Power (UT)", "scrapers.rocky_mountain", "fetch_all", "UT"),
        ("Dominion Energy Utah (UT)", "scrapers.dominion_ut", "fetch_all", "UT"),
        ("NV Energy (NV)", "scrapers.nv_energy", "fetch_all", "NV"),
        ("NorthWestern Energy (MT)", "scrapers.northwestern", "fetch_all", "MT"),
        ("Idaho Power (ID)", "scrapers.idaho_power", "fetch_all", "ID"),
        ("Avista (ID)", "scrapers.avista", "fetch_all", "ID"),
        ("DSIRE", "scrapers.dsire", "fetch_all", None),
    ]

    for label, module_path, func_name, state in sources:
        if state is not None and state not in ENABLED_STATES:
            continue
        try:
            mod = importlib.import_module(module_path)
            func = getattr(mod, func_name)
            # Multi-state sources accept the enabled-state list; single-state don't.
            result = func(ENABLED_STATES) if state is None else func()
            rows.extend(result)
        except Exception as exc:
            print("  [WARN] " + label + " failed: " + str(exc))

    # Safety net: keep only enabled states (covers multi-state sources and any
    # stray records), so ENABLED_STATES is the single source of truth for scope.
    rows = [r for r in rows if r.get("State") in ENABLED_STATES]
    return rows


def _deduplicate(rows):
    seen = {}
    for row in rows:
        key = (row.get("State", ""), _normalize(row.get("Program Name", "")))
        existing = seen.get(key)
        if existing is None:
            seen[key] = row
        else:
            # Prefer the record with more detail content
            new_detail = len(row.get("_implementation", "")) + len(row.get("_example", ""))
            old_detail = len(existing.get("_implementation", "")) + len(existing.get("_example", ""))
            if new_detail > old_detail:
                seen[key] = row
    return list(seen.values())


def _normalize(s):
    return " ".join(s.lower().split())


# Statuses that should never be auto-expired (manually set by scraper logic)
_SKIP_EXPIRE = {"Temporarily Paused", "Pending"}


def _auto_expire(rows):
    """
    Automatically set Status='Expired' for any row whose Expiration Date is a
    parseable ISO date that has already passed today.  Rows marked Temporarily
    Paused or Pending are left alone -- the scraper controls those.
    Also appends a note so users know the program lapsed and may have renewed.
    """
    today = date.today()
    changed = 0
    for row in rows:
        if row.get("Status") in _SKIP_EXPIRE:
            continue
        exp = str(row.get("Expiration Date") or "")
        try:
            exp_date = date.fromisoformat(exp)
        except ValueError:
            continue  # "Ongoing", "Pending launch", blank -- skip
        if exp_date < today and row.get("Status") != "Expired":
            row["Status"] = "Expired"
            existing_note = row.get("Notes", "")
            lapse_note = "EXPIRED " + exp + " -- program may have renewed; verify at administrator website before advising clients."
            row["Notes"] = (lapse_note + "  " + existing_note).strip() if existing_note else lapse_note
            changed += 1
    if changed:
        print("  Auto-expired " + str(changed) + " programs whose expiration date has passed.")
    return rows


def _print_summary(rows):
    counts = Counter(r["State"] for r in rows)
    for state in ENABLED_STATES:
        print("  " + state + ": " + str(counts.get(state, 0)) + " programs")
    statuses = Counter(r.get("Status", "Active") for r in rows)
    for status, n in statuses.items():
        print("  Status=" + status + ": " + str(n))


# -- SQLite ------------------------------------------------------------------

def _write_sqlite(rows):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    col_defs = ", ".join('"' + c + '" TEXT' for c in ALL_COLUMNS)
    cur.execute(
        "CREATE TABLE IF NOT EXISTS incentives "
        "(id INTEGER PRIMARY KEY AUTOINCREMENT, " + col_defs + ")"
    )
    # Migrate: add any columns introduced since the DB was first created
    existing_cols = {r[1] for r in cur.execute("PRAGMA table_info(incentives)")}
    for c in ALL_COLUMNS:
        if c not in existing_cols:
            cur.execute('ALTER TABLE incentives ADD COLUMN "' + c + '" TEXT')
    cur.execute(
        'DELETE FROM incentives WHERE "Last Scraped" = ?',
        (date.today().isoformat(),)
    )
    col_names = ", ".join('"' + c + '"' for c in ALL_COLUMNS)
    placeholders = ", ".join(["?"] * len(ALL_COLUMNS))
    insert_sql = "INSERT INTO incentives (" + col_names + ") VALUES (" + placeholders + ")"
    for row in rows:
        values = [row.get(c, "") for c in ALL_COLUMNS]
        cur.execute(insert_sql, values)
    conn.commit()
    conn.close()
    print("\nSQLite: " + str(len(rows)) + " rows written -> " + DB_PATH.name)


# -- Excel -------------------------------------------------------------------

def _write_excel(rows):
    df_all = pd.DataFrame(rows, columns=ALL_COLUMNS)
    df_all = df_all.sort_values(["State", "Incentive Type", "Program Name"])

    # Summary df (no detail columns)
    df_summary = df_all[COLUMNS]

    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="All States", index=False)
        for state in ENABLED_STATES:
            df_state = df_summary[df_summary["State"] == state]
            if not df_state.empty:
                df_state.to_excel(writer, sheet_name=state, index=False)
        # Details sheet: Program Name + structured calc values + long-form detail columns
        df_details = df_all[[
            "State", "Program Name",
            "_incentive_rate", "_rebate_tiers", "_unit_cap", "_baseline", "_min_project",
            "_implementation", "_methodology", "_example",
        ]].copy()
        df_details.columns = [
            "State", "Program Name",
            "Incentive Rate", "Rebate Tiers", "Per-Unit Cap", "Baseline Assumption", "Min Project Size",
            "Implementation Steps", "Savings Methodology", "Worked Example",
        ]
        df_details.to_excel(writer, sheet_name="Details", index=False)

    _style_excel(df_summary, df_all)
    print("Excel: " + str(len(rows)) + " rows written -> " + XLSX_PATH.name)


def _style_excel(df_summary, df_all):
    wb = load_workbook(XLSX_PATH)
    today = date.today()
    warn_date = today + timedelta(days=30)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_details = sheet_name == "Details"

        if is_details:
            # Column widths for details sheet (State, Program, 5 structured, 3 long-form)
            col_widths = [8, 38, 26, 30, 16, 30, 20, 55, 55, 55]
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            hdr_fill = PatternFill("solid", fgColor="FF1F3864")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = Font(bold=True, color="FFFFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 30
            ws.freeze_panes = "A2"
            for row_idx in range(2, ws.max_row + 1):
                bg = "FFFFFFFF" if row_idx % 2 == 0 else "FFF5F5F5"
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row_idx, col_idx)
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.font = Font(size=9)
                ws.row_dimensions[row_idx].height = 80
            continue

        df = df_summary if sheet_name == "All States" else df_summary[df_summary["State"] == sheet_name]

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            vals = [len(str(df.iloc[i][col_name])) for i in range(min(len(df), 100))]
            max_len = max([len(str(col_name))] + vals) if vals else len(str(col_name))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

        hdr_fill = PatternFill("solid", fgColor="FF1F3864")
        hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        state_col = COLUMNS.index("State") + 1
        status_col = COLUMNS.index("Status") + 1
        exp_col = COLUMNS.index("Expiration Date") + 1

        for row_idx in range(2, ws.max_row + 1):
            state_val = str(ws.cell(row_idx, state_col).value or "")
            status_val = str(ws.cell(row_idx, status_col).value or "Active")
            exp_val = str(ws.cell(row_idx, exp_col).value or "")

            if status_val == "Expired":
                bg, txt_color = "FFD3D3D3", "FF808080"
            elif status_val in ("Temporarily Paused", "Pending"):
                bg, txt_color = "FFFFF2CC", "FF000000"
            else:
                try:
                    bg = "FFFFF2CC" if date.fromisoformat(exp_val) <= warn_date else (
                        "FFFFFFFF" if row_idx % 2 == 0 else "FFF5F5F5"
                    )
                except ValueError:
                    bg = "FFFFFFFF" if row_idx % 2 == 0 else "FFF5F5F5"
                txt_color = "FF000000"

            row_fill = PatternFill("solid", fgColor=bg)
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=9, color=txt_color)

            color = STATE_COLORS.get(state_val, "FFCCCCCC")
            ws.cell(row_idx, state_col).fill = PatternFill("solid", fgColor=color)
            ws.cell(row_idx, state_col).font = Font(bold=True, color="FFFFFFFF", size=9)

    wb.save(XLSX_PATH)


# -- HTML --------------------------------------------------------------------

def _esc(s):
    """HTML-escape a string for safe embedding."""
    return (str(s or "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#39;"))


def _write_html(rows):
    df = pd.DataFrame(rows, columns=ALL_COLUMNS)
    df = df.sort_values(["State", "Status", "Program Name"])

    types = sorted(df["Incentive Type"].dropna().unique())
    sectors = sorted(df["Sector"].dropna().unique())
    techs = sorted(df["Technology"].dropna().unique())

    today = date.today()
    warn_date = today + timedelta(days=30)
    logo_uri = _logo_data_uri()

    html_colors = {
        "UT": "#4E79A7", "MT": "#59A14F",
        "ID": "#F28E2B", "NV": "#E15759",
    }

    # Scope labels derived from ENABLED_STATES (drives title, subtitle, legend).
    enabled = [s for s in ALL_STATES if s in ENABLED_STATES]
    enabled_codes = ", ".join(enabled)
    enabled_names = " &middot; ".join(STATE_NAMES.get(s, s) for s in enabled)
    # State filter + state legend chips only make sense with more than one state.
    multi_state = len(enabled) > 1
    legend_states_html = "".join(
        '<span style="background:' + html_colors.get(s, "#888") + '"></span>' + s
        for s in enabled
    ) if multi_state else ""
    if multi_state:
        _state_opts = '<option value="">All</option>' + "".join(
            "<option>" + _esc(s) + "</option>" for s in enabled)
        state_filter_html = '<select id="f-state" onchange="applyFilters()">' + _state_opts + "</select>"
    else:
        state_filter_html = ""

    # Build per-row detail JSON (indexed by row id)
    detail_data = {}
    rows_html = []

    for row_idx, (_, row) in enumerate(df.iterrows()):
        status = str(row.get("Status") or "Active")
        exp = str(row.get("Expiration Date") or "")
        state = str(row.get("State") or "")
        color = html_colors.get(state, "#888")

        row_class = ""
        if status == "Expired":
            row_class = "expired"
        elif status in ("Temporarily Paused", "Pending"):
            row_class = "paused"
        else:
            try:
                if date.fromisoformat(exp) <= warn_date:
                    row_class = "expiring"
            except ValueError:
                pass

        url = str(row.get("Application URL") or "")
        name = str(row.get("Program Name") or "")

        # Auto-classify this program into UCREW equipment categories by scanning
        # its full text. Powers the AR Finder search and the modal chips.
        classify_text = " ".join(str(row.get(c) or "") for c in (
            "Program Name", "Technology", "Sector", "Incentive Type",
            "_methodology", "_implementation", "_example", "Notes",
        ))
        equipment = classify_equipment(classify_text)
        is_custom = any(c in CUSTOM_CATEGORIES for c in equipment)

        # Store detail data in JS object
        detail_data[row_idx] = {
            "name": name,
            "state": state,
            "admin": str(row.get("Administrator") or ""),
            "sector": str(row.get("Sector") or ""),
            "type": str(row.get("Incentive Type") or ""),
            "tech": str(row.get("Technology") or ""),
            "value": str(row.get("Incentive Value") or ""),
            "max": str(row.get("Max Benefit") or ""),
            "recipients": str(row.get("Eligible Recipients") or ""),
            "expiration": exp,
            "url": url,
            "status": status,
            "notes": str(row.get("Notes") or ""),
            "implementation": str(row.get("_implementation") or ""),
            "methodology": str(row.get("_methodology") or ""),
            "example": str(row.get("_example") or ""),
            "lastScraped": str(row.get("Last Scraped") or ""),
            # Structured calculation values
            "rate": str(row.get("_incentive_rate") or ""),
            "tiers": str(row.get("_rebate_tiers") or ""),
            "cap": str(row.get("_unit_cap") or ""),
            "baseline": str(row.get("_baseline") or ""),
            "minProject": str(row.get("_min_project") or ""),
            "equipment": equipment,
            "custom": is_custom,
        }

        cells = [
            '<td><span class="badge" style="background:' + color + '">' + _esc(state) + "</span></td>",
            '<td class="name-cell"><button class="detail-btn" onclick="openDetail(' + str(row_idx) + ')">'
            + _esc(name) + "</button></td>",
        ]
        for col in COLUMNS[2:]:
            val = str(row.get(col) or "")
            if col == "Application URL" and val:
                cells.append('<td><a href="' + _esc(val) + '" target="_blank">Open</a></td>')
            elif col in ("Incentive Value", "Max Benefit"):
                # Headline dollar figures -- emphasized so they are scannable in the table
                cells.append('<td class="money">' + _esc(val) + "</td>")
            else:
                cells.append("<td>" + _esc(val) + "</td>")

        rows_html.append(
            '<tr class="' + row_class + '" '
            'data-state="' + _esc(state) + '" '
            'data-type="' + _esc(str(row.get("Incentive Type") or "")) + '" '
            'data-sector="' + _esc(str(row.get("Sector") or "")) + '" '
            'data-tech="' + _esc(str(row.get("Technology") or "")) + '" '
            'data-equipment="' + _esc(";".join(equipment)) + '" '
            'data-custom="' + ("1" if is_custom else "0") + '" '
            'data-idx="' + str(row_idx) + '">'
            + "".join(cells) + "</tr>"
        )

    def options(vals):
        return '<option value="">All</option>' + "".join(
            "<option>" + _esc(v) + "</option>" for v in vals
        )

    th_cells = "".join(
        '<th onclick="sortTable(' + str(i) + ')">' + col + "</th>"
        for i, col in enumerate(COLUMNS)
    )

    detail_json = json.dumps(detail_data, ensure_ascii=True)

    html = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>UCREW -- Commercial &amp; Industrial Energy Incentives (""" + enabled_codes + """)</title>
<link rel="icon" type="image/svg+xml" href=\"""" + logo_uri + """\">
<style>
:root {
  --brand: """ + BRAND + """;
  --brand-dark: """ + BRAND_DARK + """;
  --brand-darker: """ + BRAND_DARKER + """;
  --brand-tint: #fbeaea;          /* light red row/hover wash */
  --brand-tint-border: #f0cccc;
  --money: #1f6e1f;               /* green -- dollar figures / savings */
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; font-size: 13px; background: #f4f6f9; color: #222; }
header { background: linear-gradient(135deg, var(--brand) 0%, var(--brand-dark) 100%); color: #fff; padding: 18px 24px; }
.header-inner { display: flex; align-items: center; gap: 18px; }
.brand-logo { height: 46px; width: auto; flex-shrink: 0; filter: brightness(0) invert(1); }
.header-text { flex: 1; min-width: 0; }
header h1 { font-size: 20px; }
header p { font-size: 12px; opacity: .8; margin-top: 4px; }
.controls { padding: 12px 24px; background: #fff; border-bottom: 1px solid #ddd; display: flex; flex-wrap: wrap; gap: 10px; align-items: center; }
.controls input, .controls select { padding: 6px 10px; border: 1px solid #ccc; border-radius: 4px; font-size: 12px; }
.controls input { width: 200px; }
.legend { margin-left: auto; display: flex; gap: 8px; align-items: center; font-size: 11px; flex-wrap: wrap; }
.legend span { display: inline-block; width: 12px; height: 12px; border-radius: 2px; }
#count { font-size: 12px; color: #555; }
.table-wrap { overflow-x: auto; padding: 16px 24px; }
table { width: 100%; border-collapse: collapse; background: #fff; box-shadow: 0 1px 4px rgba(0,0,0,.08); border-radius: 6px; overflow: hidden; }
th { background: var(--brand); color: #fff; padding: 8px 10px; text-align: left; cursor: pointer; white-space: nowrap; font-size: 12px; user-select: none; }
th:hover { background: var(--brand-dark); }
th.asc::after { content: " \25b2"; font-size: 10px; }
th.desc::after { content: " \25bc"; font-size: 10px; }
td { padding: 6px 10px; border-bottom: 1px solid #eee; vertical-align: top; font-size: 12px; }
tr:last-child td { border-bottom: none; }
tr:hover td { background: var(--brand-tint) !important; }
tr.expired td { color: #aaa; background: #fafafa; }
tr.paused td { background: #fffde7; }
tr.expiring td { background: #fff8e1; }
tr.hidden { display: none; }
.badge { display: inline-block; padding: 2px 8px; border-radius: 10px; color: #fff; font-weight: 700; font-size: 11px; }
a { color: var(--brand); text-decoration: none; }
a:hover { text-decoration: underline; }
.detail-btn { background: none; border: none; color: var(--brand); cursor: pointer; text-align: left; font-size: 12px; padding: 0; font-family: inherit; text-decoration: underline; text-decoration-style: dotted; }
.detail-btn:hover { color: var(--brand-dark); }
td:nth-child(2) { min-width: 180px; max-width: 280px; }
td.money { font-weight: 700; color: var(--money); white-space: nowrap; }
tr.expired td.money { color: #9bb69b; }
.header-badges { margin-top: 8px; display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }
.hbadge { display: inline-block; background: rgba(255,255,255,.14); border: 1px solid rgba(255,255,255,.25); color: #fff; font-size: 12px; padding: 3px 10px; border-radius: 12px; }
.hbadge.scanned { background: #2d8c2d; border-color: #2d8c2d; font-weight: 600; }
.dl-links { display: flex; gap: 8px; align-items: center; }
.dl-btn { display: inline-block; background: var(--brand-tint); border: 1px solid var(--brand-tint-border); color: var(--brand); padding: 5px 10px; border-radius: 4px; font-size: 12px; font-weight: 600; }
.dl-btn:hover { background: #f6dada; text-decoration: none; }

/* AR Finder */
.ar-bar { background: linear-gradient(135deg, #fff 0%, var(--brand-tint) 100%); border-bottom: 1px solid var(--brand-tint-border); padding: 16px 24px; }
.ar-inner { max-width: 960px; }
.ar-label { display: block; font-size: 13px; font-weight: 700; color: var(--brand); margin-bottom: 8px; letter-spacing: .2px; }
.ar-row { display: flex; gap: 8px; }
.ar-input { flex: 1; padding: 11px 14px; border: 2px solid var(--brand-tint-border); border-radius: 6px; font-size: 14px; font-family: inherit; background: #fff; }
.ar-input:focus { outline: none; border-color: var(--brand); box-shadow: 0 0 0 3px rgba(190,0,0,.12); }
.ar-clear { border: 1px solid #ccc; background: #fff; color: #555; border-radius: 6px; padding: 0 14px; font-size: 12px; font-weight: 600; cursor: pointer; }
.ar-clear:hover { background: #f4f4f4; }
.ar-result { font-size: 12.5px; color: #333; margin-top: 10px; line-height: 1.6; }
.ar-result:empty { display: none; }
.ar-result .eq-chip { display: inline-block; background: var(--brand); color: #fff; border-radius: 11px; padding: 2px 9px; font-size: 11px; font-weight: 700; margin: 0 3px 3px 0; }
.ar-result .miss { color: #999; }
.ar-result b { color: var(--brand); }
.ar-examples { margin-top: 10px; font-size: 12px; color: #777; display: flex; flex-wrap: wrap; gap: 6px; align-items: center; }
.ar-try { font-weight: 600; }
.ar-chip-btn { border: 1px solid var(--brand-tint-border); background: #fff; color: var(--brand); border-radius: 12px; padding: 3px 10px; font-size: 11.5px; font-family: inherit; cursor: pointer; }
.ar-chip-btn:hover { background: var(--brand-tint); }
tr.ar-custom td { background: #fffdf5; }
tr.ar-custom td:first-child { box-shadow: inset 3px 0 0 #e6b800; }
td .eq-tag { display: inline-block; background: #f0e6e6; color: #7a3a3a; border-radius: 9px; padding: 1px 7px; font-size: 10px; font-weight: 700; margin: 1px 2px 1px 0; }
.match-tag { display: inline-block; background: var(--brand); color: #fff; border-radius: 9px; padding: 1px 7px; font-size: 10px; font-weight: 700; margin-left: 6px; vertical-align: middle; }
.custom-tag { display: inline-block; background: #e6b800; color: #4a3b00; border-radius: 9px; padding: 1px 7px; font-size: 10px; font-weight: 700; margin-left: 6px; vertical-align: middle; }
.equip-chip { display: inline-block; background: var(--brand-tint); color: var(--brand); border: 1px solid var(--brand-tint-border); border-radius: 12px; padding: 4px 11px; font-size: 12px; font-weight: 700; margin: 0 5px 5px 0; }

/* Modal */
.overlay { display: none; position: fixed; inset: 0; background: rgba(0,0,0,.55); z-index: 100; overflow-y: auto; padding: 40px 20px; }
.overlay.open { display: flex; align-items: flex-start; justify-content: center; }
.modal { background: #fff; border-radius: 10px; max-width: 800px; width: 100%; box-shadow: 0 8px 40px rgba(0,0,0,.25); }
.modal-header { background: var(--brand); color: #fff; padding: 20px 24px; border-radius: 10px 10px 0 0; display: flex; align-items: flex-start; gap: 12px; }
.modal-header h2 { font-size: 16px; flex: 1; line-height: 1.4; }
.modal-header .state-badge { font-size: 13px; font-weight: 700; padding: 3px 10px; border-radius: 12px; white-space: nowrap; }
.close-btn { background: none; border: none; color: rgba(255,255,255,.8); font-size: 22px; cursor: pointer; padding: 0 4px; line-height: 1; }
.close-btn:hover { color: #fff; }
.modal-body { padding: 0; }
.meta-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 0; border-bottom: 1px solid #eee; }
.meta-item { padding: 12px 16px; border-right: 1px solid #eee; border-bottom: 1px solid #eee; }
.meta-item:nth-child(3n) { border-right: none; }
.meta-label { font-size: 10px; text-transform: uppercase; letter-spacing: .5px; color: #888; margin-bottom: 3px; }
.meta-value { font-size: 13px; font-weight: 600; color: #222; }
.meta-value.highlight { color: #1f6e1f; font-size: 15px; }
.meta-value a { color: var(--brand); }
.status-active { color: #1f6e1f; }
.status-paused { color: #b45309; }
.status-pending { color: #6d4c41; }
.status-expired { color: #888; }
.section { padding: 20px 24px; border-bottom: 1px solid #eee; }
.section:last-child { border-bottom: none; }
.section-title { font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: .5px; color: var(--brand); margin-bottom: 10px; display: flex; align-items: center; gap: 8px; }
.section-title::after { content: ''; flex: 1; height: 1px; background: #e0e7f0; }
.section-body { font-size: 13px; line-height: 1.7; color: #333; white-space: pre-line; }
.example-box { background: #f0f7f0; border-left: 3px solid #2d8c2d; border-radius: 0 6px 6px 0; padding: 14px 16px; font-size: 13px; line-height: 1.7; white-space: pre-line; }
.notes-box { background: #fff8e1; border-left: 3px solid #e6b800; border-radius: 0 6px 6px 0; padding: 12px 16px; font-size: 12px; color: #555; }
.calc-panel { background: #f0f7f0; border: 1px solid #cfe6cf; border-radius: 8px; padding: 4px 0; }
.calc-row { display: grid; grid-template-columns: 190px 1fr; gap: 10px; padding: 9px 16px; border-bottom: 1px solid #dcecdc; }
.calc-row:last-child { border-bottom: none; }
.calc-label { font-size: 11px; text-transform: uppercase; letter-spacing: .4px; color: #4a7a4a; font-weight: 700; align-self: center; }
.calc-val { font-size: 15px; font-weight: 700; color: #1f6e1f; }
.num { background: #d9f0d9; color: #145214; border-radius: 3px; padding: 0 3px; font-weight: 700; }
@media (max-width: 520px) { .calc-row { grid-template-columns: 1fr; gap: 2px; } }
.modal-footer { padding: 16px 24px; background: #f8f9fa; border-radius: 0 0 10px 10px; display: flex; justify-content: space-between; align-items: center; }
.apply-btn { display: inline-block; background: var(--brand); color: #fff; padding: 9px 20px; border-radius: 5px; text-decoration: none; font-size: 13px; font-weight: 600; }
.apply-btn:hover { background: var(--brand-dark); color: #fff; text-decoration: none; }
.last-scraped { font-size: 11px; color: #999; }
</style>
</head>
<body>
<header>
  <div class="header-inner">
    <img class="brand-logo" src=\"""" + logo_uri + """\" alt="UCREW">
    <div class="header-text">
      <h1>Commercial &amp; Industrial Energy Incentives</h1>
      <p>""" + enabled_names + """ &nbsp;|&nbsp; Click any program name for the full calculation breakdown</p>
      <div class="header-badges">
        <span class="hbadge scanned">Last scanned: """ + str(today) + """ &middot; refreshes daily</span>
        <span class="hbadge">""" + str(len(rows)) + """ programs</span>
        <span class="hbadge">Commercial &amp; Industrial facilities only</span>
      </div>
    </div>
  </div>
</header>
<div class="ar-bar">
  <div class="ar-inner">
    <label class="ar-label" for="ar-input">&#128269; Find incentives for a recommendation (AR)</label>
    <div class="ar-row">
      <input type="text" id="ar-input" class="ar-input" autocomplete="off"
             placeholder="Describe the measure, e.g. &quot;Install VFD on compressors&quot; or &quot;Install lighting controls&quot;"
             oninput="applyFilters()">
      <button type="button" class="ar-clear" id="ar-clear" onclick="clearAr()">Clear</button>
    </div>
    <div class="ar-result" id="ar-result"></div>
    <div class="ar-examples"><span class="ar-try">Try:</span>
      <button type="button" class="ar-chip-btn" onclick="setAr(this.textContent)">Install VFD on compressors</button>
      <button type="button" class="ar-chip-btn" onclick="setAr(this.textContent)">Install lighting controls</button>
      <button type="button" class="ar-chip-btn" onclick="setAr(this.textContent)">Add VFD to HVAC supply fan</button>
      <button type="button" class="ar-chip-btn" onclick="setAr(this.textContent)">Replace boiler; add heat recovery</button>
      <button type="button" class="ar-chip-btn" onclick="setAr(this.textContent)">Refrigeration controls upgrade</button>
    </div>
  </div>
</div>
<div class="controls">
  <input type="text" id="search" placeholder="Search all fields..." oninput="applyFilters()">
  """ + state_filter_html + """
  <select id="f-type" onchange="applyFilters()"><option value="">All Types</option>""" + options(types) + """</select>
  <select id="f-sector" onchange="applyFilters()"><option value="">All Sectors</option>""" + options(sectors) + """</select>
  <select id="f-tech" onchange="applyFilters()"><option value="">All Technologies</option>""" + options(techs) + """</select>
  <select id="f-status" onchange="applyFilters()">
    <option value="">All Statuses</option>
    <option>Active</option><option>Temporarily Paused</option><option>Pending</option><option>Expired</option>
  </select>
  <span id="count"></span>
  <div class="dl-links">
    <a class="dl-btn" href="incentives.xlsx" download>&#8681; Excel</a>
    <a class="dl-btn" href="index.html" download="ucrew-incentives.html">&#8681; This page</a>
  </div>
  <div class="legend">
    """ + legend_states_html + """
    <span style="background:#fffde7;border:1px solid #ccc"></span>Paused/Pending
    <span style="background:#fff8e1;border:1px solid #ccc"></span>Expiring Soon
  </div>
</div>
<div class="table-wrap">
<table id="tbl">
<thead><tr>""" + th_cells + """</tr></thead>
<tbody>
""" + "\n".join(rows_html) + """
</tbody>
</table>
</div>

<!-- Detail Modal -->
<div class="overlay" id="overlay" onclick="closeOnOverlay(event)">
<div class="modal" id="modal">
  <div class="modal-header">
    <h2 id="m-name"></h2>
    <span class="state-badge" id="m-state-badge"></span>
    <button class="close-btn" onclick="closeDetail()">&#10005;</button>
  </div>
  <div class="modal-body">
    <div class="meta-grid">
      <div class="meta-item"><div class="meta-label">Incentive Value</div><div class="meta-value highlight" id="m-value"></div></div>
      <div class="meta-item"><div class="meta-label">Max Benefit</div><div class="meta-value" id="m-max"></div></div>
      <div class="meta-item"><div class="meta-label">Status</div><div class="meta-value" id="m-status"></div></div>
      <div class="meta-item"><div class="meta-label">Incentive Type</div><div class="meta-value" id="m-type"></div></div>
      <div class="meta-item"><div class="meta-label">Technology</div><div class="meta-value" id="m-tech"></div></div>
      <div class="meta-item"><div class="meta-label">Sector</div><div class="meta-value" id="m-sector"></div></div>
      <div class="meta-item"><div class="meta-label">Administrator</div><div class="meta-value" id="m-admin"></div></div>
      <div class="meta-item"><div class="meta-label">Eligible Recipients</div><div class="meta-value" id="m-recip"></div></div>
      <div class="meta-item"><div class="meta-label">Expiration</div><div class="meta-value" id="m-exp"></div></div>
    </div>
    <div id="m-equip-section" class="section">
      <div class="section-title">Applies to Equipment</div>
      <div id="m-equip"></div>
    </div>
    <div id="m-calc-section" class="section">
      <div class="section-title">Calculation Values (numbers used in savings math)</div>
      <div class="calc-panel" id="m-calc-panel"></div>
    </div>
    <div id="m-notes-section" class="section">
      <div class="section-title">Important Notes</div>
      <div class="notes-box" id="m-notes"></div>
    </div>
    <div class="section">
      <div class="section-title">How to Apply (Implementation Steps)</div>
      <div class="section-body" id="m-implementation"></div>
    </div>
    <div class="section">
      <div class="section-title">Savings Methodology</div>
      <div class="section-body" id="m-methodology"></div>
    </div>
    <div class="section">
      <div class="section-title">Worked Example</div>
      <div class="example-box" id="m-example"></div>
    </div>
  </div>
  <div class="modal-footer">
    <a class="apply-btn" id="m-apply-link" href="#" target="_blank">Apply / More Info &rarr;</a>
    <span class="last-scraped" id="m-last-scraped"></span>
  </div>
</div>
</div>

<script>
var DETAILS = """ + detail_json + """;
var HTML_COLORS = {"UT":"#4E79A7","MT":"#59A14F","ID":"#F28E2B","NV":"#E15759"};
// Equipment vocabulary -- same map the build step used to tag each program, so a
// typed recommendation resolves to the exact categories the programs carry.
var EQUIP_KEYWORDS = """ + json.dumps(EQUIPMENT_KEYWORDS, ensure_ascii=True) + """;
var CUSTOM_CATEGORIES = """ + json.dumps(sorted(CUSTOM_CATEGORIES), ensure_ascii=True) + """;
var sortCol = -1, sortDir = 1;

// Word-boundary keyword test mirroring the Python classifier: the keyword must
// start at a word boundary, but trailing letters are allowed ("compressor"
// matches "compressors"). Avoids false hits like "led" inside "controlled".
function kwHit(hay, needle) {
  var idx = hay.indexOf(needle);
  while (idx >= 0) {
    var before = idx === 0 ? ' ' : hay.charAt(idx - 1);
    if (!/[a-z0-9]/.test(before)) return true;
    idx = hay.indexOf(needle, idx + 1);
  }
  return false;
}

// Map a free-text assessment recommendation to equipment categories by scanning
// for the same keywords used to tag programs. Custom/whole-facility buckets are
// excluded here -- those surface automatically for every AR, not by keyword.
function arCategories(text) {
  var t = String(text).toLowerCase();
  var cats = [];
  for (var cat in EQUIP_KEYWORDS) {
    if (CUSTOM_CATEGORIES.indexOf(cat) >= 0) continue;
    var kws = EQUIP_KEYWORDS[cat];
    for (var i = 0; i < kws.length; i++) {
      if (kwHit(t, kws[i])) { cats.push(cat); break; }
    }
  }
  return cats;
}

function setAr(text) {
  document.getElementById('ar-input').value = text;
  applyFilters();
  document.getElementById('ar-input').focus();
}

function clearAr() {
  document.getElementById('ar-input').value = '';
  applyFilters();
}

// Escape HTML, then wrap money / % / kWh / therm figures in a highlight span so the
// actual numbers used in calculations stand out inside the prose sections.
function hlNums(text) {
  var esc = String(text)
    .replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
  return esc.replace(
    /(\\$[\\d,]+(?:\\.\\d+)?(?:\\/[A-Za-z²]+)?|\\b\\d[\\d,]*(?:\\.\\d+)?\\s?(?:%|kWh|kW|MWh|therms?|sq ?ft|SEER2?|HSPF2?|EER|AFUE|COP|UEF|HDD|CDD)\\b)/g,
    '<span class="num">$1</span>'
  );
}

function openDetail(idx) {
  var d = DETAILS[idx];
  if (!d) return;
  document.getElementById('m-name').textContent = d.name;
  var badge = document.getElementById('m-state-badge');
  badge.textContent = d.state;
  badge.style.background = HTML_COLORS[d.state] || '#888';
  document.getElementById('m-value').textContent = d.value || 'See program';
  document.getElementById('m-max').textContent = d.max || 'See program';
  var statusEl = document.getElementById('m-status');
  statusEl.textContent = d.status;
  statusEl.className = 'meta-value status-' + d.status.toLowerCase().replace(/\\s+/g, '-').replace('temporarily-', '');
  document.getElementById('m-type').textContent = d.type;
  document.getElementById('m-tech').textContent = d.tech;
  document.getElementById('m-sector').textContent = d.sector;
  document.getElementById('m-admin').textContent = d.admin;
  document.getElementById('m-recip').textContent = d.recipients;
  document.getElementById('m-exp').textContent = d.expiration || 'Ongoing';
  var notesSection = document.getElementById('m-notes-section');
  var notesEl = document.getElementById('m-notes');
  if (d.notes) {
    notesEl.textContent = d.notes;
    notesSection.style.display = '';
  } else {
    notesSection.style.display = 'none';
  }
  // Equipment categories this program applies to (auto-classified)
  var equipSection = document.getElementById('m-equip-section');
  var equip = d.equipment || [];
  if (equip.length) {
    document.getElementById('m-equip').innerHTML = equip.map(function(c) {
      return '<span class="equip-chip">' + escHtml(c) + '</span>';
    }).join(' ');
    equipSection.style.display = '';
  } else {
    equipSection.style.display = 'none';
  }
  // Calculation Values panel -- only rows with a value are shown; hide panel if all blank
  var calcFields = [
    ['Incentive rate', d.rate],
    ['Rebate tiers', d.tiers],
    ['Per-unit / project cap', d.cap],
    ['Baseline assumption', d.baseline],
    ['Minimum project size', d.minProject],
  ];
  var calcPanel = document.getElementById('m-calc-panel');
  var calcHtml = '';
  for (var i = 0; i < calcFields.length; i++) {
    var label = calcFields[i][0], val = calcFields[i][1];
    if (val) {
      calcHtml += '<div class="calc-row"><div class="calc-label">' + label +
        '</div><div class="calc-val">' + hlNums(val) + '</div></div>';
    }
  }
  document.getElementById('m-calc-section').style.display = calcHtml ? '' : 'none';
  calcPanel.innerHTML = calcHtml;

  document.getElementById('m-implementation').textContent = d.implementation || 'See program website for application instructions.';
  document.getElementById('m-methodology').innerHTML = d.methodology ? hlNums(d.methodology) : 'See program website for savings calculation methodology.';
  document.getElementById('m-example').innerHTML = d.example ? hlNums(d.example) : 'See program website for example savings calculations.';
  var applyLink = document.getElementById('m-apply-link');
  applyLink.href = d.url || '#';
  applyLink.style.display = d.url ? '' : 'none';
  document.getElementById('m-last-scraped').textContent = 'Data last verified: ' + d.lastScraped;
  document.getElementById('overlay').classList.add('open');
  document.body.style.overflow = 'hidden';
}

function closeDetail() {
  document.getElementById('overlay').classList.remove('open');
  document.body.style.overflow = '';
}

function closeOnOverlay(e) {
  if (e.target === document.getElementById('overlay')) closeDetail();
}

document.addEventListener('keydown', function(e) {
  if (e.key === 'Escape') closeDetail();
});

function escHtml(s) {
  return String(s).replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
}

// Remove any AR match tag / styling from a row (called every pass so tags never stack).
function annotateRow(row, arActive, overlap, isCustom) {
  var cell = row.cells[1];
  var old = cell.querySelector('.match-tag, .custom-tag');
  if (old) old.remove();
  row.classList.remove('ar-custom');
  if (!arActive) return;
  if (overlap.length > 0) {
    var s = document.createElement('span');
    s.className = 'match-tag';
    s.textContent = 'match: ' + overlap.join(', ');
    cell.appendChild(s);
  } else if (isCustom) {
    var c = document.createElement('span');
    c.className = 'custom-tag';
    c.textContent = 'custom / whole-facility';
    cell.appendChild(c);
    row.classList.add('ar-custom');
  }
}

// When an AR is active, float the best matches to the top: specific equipment
// matches first (most overlap first), then custom/whole-facility, expired last.
function reorderForAr() {
  var tbody = document.querySelector('#tbl tbody');
  var rows = Array.from(tbody.querySelectorAll('tr'));
  rows.sort(function(a, b) {
    function grp(r) { return r._arSpecific ? 0 : (r._arCustom ? 1 : 2); }
    var ga = grp(a), gb = grp(b);
    if (ga !== gb) return ga - gb;
    if (a._expired !== b._expired) return a._expired ? 1 : -1;
    if (b._arScore !== a._arScore) return b._arScore - a._arScore;
    return a.cells[1].textContent.localeCompare(b.cells[1].textContent);
  });
  rows.forEach(function(r) { tbody.appendChild(r); });
  document.querySelectorAll('th').forEach(function(th) { th.className = ''; });
  sortCol = -1;
}

function renderArBanner(arActive, arText, arCats, specificCount, customCount) {
  var el = document.getElementById('ar-result');
  if (!arActive) { el.innerHTML = ''; return; }
  if (arCats.length === 0) {
    el.innerHTML = 'No specific equipment recognized in <b>&ldquo;' + escHtml(arText) +
      '&rdquo;</b>. Showing custom / whole-facility programs that can cover most measures, plus any text matches. ' +
      'Try naming the equipment &mdash; compressor, lighting, boiler, motor/VFD, fan, pump, refrigeration&hellip;';
    return;
  }
  var chips = arCats.map(function(c) { return '<span class="eq-chip">' + c + '</span>'; }).join(' ');
  var msg = 'Matched equipment: ' + chips + ' &mdash; <b>' + specificCount + '</b> targeted incentive' +
    (specificCount === 1 ? '' : 's');
  if (customCount > 0) {
    msg += ', plus <b>' + customCount + '</b> custom / whole-facility program' +
      (customCount === 1 ? '' : 's') + ' that typically also apply';
  }
  el.innerHTML = msg + '. Best matches are listed first.';
}

function applyFilters() {
  var q = document.getElementById('search').value.toLowerCase();
  var stateEl = document.getElementById('f-state');
  var state = stateEl ? stateEl.value : '';
  var type = document.getElementById('f-type').value;
  var sector = document.getElementById('f-sector').value;
  var tech = document.getElementById('f-tech').value;
  var status = document.getElementById('f-status').value;
  var arText = document.getElementById('ar-input').value.trim();
  var arActive = arText.length > 0;
  var arCats = arActive ? arCategories(arText) : [];
  document.getElementById('ar-clear').style.visibility = arActive ? 'visible' : 'hidden';

  var visible = 0, specificCount = 0, customCount = 0;
  document.querySelectorAll('#tbl tbody tr').forEach(function(row) {
    var text = row.textContent.toLowerCase();
    var idx = row.dataset.idx;
    var d = DETAILS[idx] || {};
    var fullText = text + ' ' + (d.implementation || '').toLowerCase() + ' ' + (d.example || '').toLowerCase();
    var base =
      (!q || fullText.indexOf(q) >= 0) &&
      (!state || row.dataset.state === state) &&
      (!type || row.dataset.type === type) &&
      (!sector || row.dataset.sector === sector) &&
      (!tech || row.dataset.tech === tech) &&
      (!status || (row.cells[12] && row.cells[12].textContent.trim() === status));

    var overlap = [];
    var isCustom = row.dataset.custom === '1';
    if (arActive) {
      var eq = (row.dataset.equipment || '').split(';').filter(Boolean);
      overlap = eq.filter(function(c) { return arCats.indexOf(c) >= 0; });
    }
    row._arScore = overlap.length;
    row._arSpecific = overlap.length > 0;
    row._arCustom = isCustom;
    row._expired = row.classList.contains('expired');

    var arOk = !arActive || overlap.length > 0 || isCustom;
    var match = base && arOk;
    annotateRow(row, arActive, overlap, isCustom);
    row.classList.toggle('hidden', !match);
    if (match) {
      visible++;
      if (arActive && overlap.length > 0) specificCount++;
      else if (arActive && isCustom) customCount++;
    }
  });

  if (arActive) reorderForAr();
  document.getElementById('count').textContent = visible + ' programs shown';
  renderArBanner(arActive, arText, arCats, specificCount, customCount);
}

function sortTable(col) {
  var tbody = document.querySelector('#tbl tbody');
  var rows = Array.from(tbody.querySelectorAll('tr'));
  if (sortCol === col) { sortDir *= -1; } else { sortCol = col; sortDir = 1; }
  rows.sort(function(a, b) {
    var av = a.cells[col] ? a.cells[col].textContent.trim() : '';
    var bv = b.cells[col] ? b.cells[col].textContent.trim() : '';
    return av.localeCompare(bv, undefined, {numeric: true}) * sortDir;
  });
  rows.forEach(function(r) { tbody.appendChild(r); });
  document.querySelectorAll('th').forEach(function(th, i) {
    th.className = i === col ? (sortDir === 1 ? 'asc' : 'desc') : '';
  });
}

applyFilters();
</script>
</body>
</html>"""

    HTML_PATH.write_text(html, encoding="utf-8")
    print("HTML: " + str(len(rows)) + " rows written -> " + HTML_PATH.name)


# -- Log ---------------------------------------------------------------------

def _write_log(count):
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(datetime.now().isoformat() + " -- " + str(count) + " programs fetched\n")


if __name__ == "__main__":
    main()
